from telegram import Update, ReplyKeyboardRemove, ReplyKeyboardMarkup, ReplyKeyboardRemove
from telegram.ext import ApplicationBuilder, ContextTypes, CommandHandler, filters, ConversationHandler, MessageHandler

import octb.modules.sql.product as sql
from octb.modules.sql.product import Product
from octb import LOGGER
import asyncio

from octb import application

from zipfile import ZipFile
import openpyxl
import os
import shutil

__mod_name__ = "import_bulk"

from decouple import config
SUPERADMIN_ID=int(config('SUPERADMIN_ID'))
STORAGE = config('STORAGE')

from openpyxl import Workbook
async def import_bulk(update: Update, context:ContextTypes.DEFAULT_TYPE):
    user = update.message.from_user
    user_text = update.message.text

    if user.id != SUPERADMIN_ID:
        await update.message.reply_text("У вас нет прав")
        return

    command = user_text.split("\n")[0].split(" ")
    if "-y" not in command:
        await update.message.reply_text("Вы не подтвердили отправку с помощью -y")
        return
    if len(command) != 3:
        await update.message.reply_text("/import_bulk -y USER_ID")
        return

    seller_id = int(command[2])

    message_replied = update.message.reply_to_message
    file_name = f"{seller_id}.{message_replied.document.file_name.split('.')[-1]}"
    storage_path = f"{STORAGE}/files/import/{file_name}"
    storage_folder = storage_path+"fd"


    # download zip
    if os.path.exists(storage_path):
        os.remove(storage_path)
    with open(storage_path, 'wb') as f:
        file = await context.bot.get_file(message_replied.document)
        await file.download(out=f)

    # unarchive zip
    if os.path.exists(storage_path):
        shutil.rmtree(storage_folder, ignore_errors=True)
    with ZipFile(storage_path, 'r') as zipObj:
        # Extract all the contents of zip file in different directory
        zipObj.extractall(storage_folder)
    
    # check if xslx file exists
    prices_path = storage_folder+"/prices.xlsx"
    if not os.path.exists(prices_path):
        await context.bot.send_message(chat_id=update.effective_chat.id, text="No prices.xlsx!")
        return

    # add products from xlsx
    wb_obj = openpyxl.load_workbook(prices_path)
    sheet_obj = wb_obj.active

    m_row = sheet_obj.max_row

    products = []
    for i in range(2, m_row + 1):
        print(i)
        if not sheet_obj.cell(row = i, column = 1).value:
            LOGGER.info("id not found, continuing")
            continue
        info = {
            "id": str(sheet_obj.cell(row = i, column = 1).value or '').strip(),
            "name": str(sheet_obj.cell(row = i, column = 2).value or '').strip(),
            "description": str(sheet_obj.cell(row = i, column = 3).value or '').strip(),
            "product_type": str(sheet_obj.cell(row = i, column = 4).value or '').lower().strip(),
            "price": str(sheet_obj.cell(row = i, column = 5).value or '').strip(),
            "category": str(sheet_obj.cell(row = i, column = 6).value or '').lower().strip(),
            "subcategory": str(sheet_obj.cell(row = i, column = 7).value or '').lower().strip(),
        }

        has_image = False
        if os.path.exists(f"{storage_folder}/photos/{info['id']}.jpg"):
            has_image = True

        # Product(name, has_image, description, product_type, price, seller_id, subcategory, message_id)
        category_obj = sql.get_category_by_name(info['category'])
        subcategory_obj = sql.get_subcategory_by_name_and_category_id(info['subcategory'], category_obj.id)
        print(info['name'], info['subcategory'], category_obj.name, subcategory_obj)

        product = sql.add_product_by_obj(Product(info['name'], has_image, info['description'], sql.ProductTypeEnum(info['product_type']),
                                info['price'], seller_id, subcategory_obj.id, None))

        if has_image:
            os.rename(f"{storage_folder}/photos/{info['id']}.jpg", f'{STORAGE}/photos/product/{product.id}.jpg')
        # TODO add removal of garbage files
        products.append(product)

    await context.bot.send_message(chat_id=seller_id, text=f"Added {len(products)} products!")
    await context.bot.send_message(chat_id=update.effective_chat.id, text=f"Added {len(products)} products!")

# handlers
import_bulk_handler = CommandHandler('import_bulk', import_bulk, filters=filters.ChatType.PRIVATE)

application.add_handlers([import_bulk_handler])
